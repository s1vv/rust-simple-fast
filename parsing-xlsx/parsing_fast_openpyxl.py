# type: ignore
import json
from math import ceil
from typing import Any, TextIO
from openpyxl import load_workbook


def parse_excel_file(
    file_path: str,
    log_messages: TextIO | None = None,
) -> tuple[list[dict[str, Any]], int]:
    try:
        wb = load_workbook(filename=file_path, read_only=True, data_only=True)
        sheet = wb.active
    except Exception as e:
        if log_messages:
            log_messages.write(f"Ошибка при чтении Excel: {e}\n")
        raise

    # Первая строка — заголовки
    headers_row = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True))
    headers = {col: idx for idx, col in enumerate(headers_row) if col}

    supplier_data: dict[str, dict[str, Any]] = {}
    err_count = 0

    # Перебираем начиная со 2-й строки
    for idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True)):
        name = None
        try:
            name = row[headers["Название"]]
            stock = row[headers["Остаток"]]
            price = ceil(float(row[headers["Цена за 1 уп."]]))

            if name not in supplier_data or price > supplier_data[name]["price"]:
                supplier_data[name] = {"stock": stock, "price": price}

        except Exception as e:
            err_count += 1
            if log_messages:
                msg = f"Ошибка в строке {idx + 2}"
                if name:
                    msg += f" '{name}'"
                msg += f": {e}\n"
                log_messages.write(msg)

    data_list = [
        {"name": name, "stock": data["stock"], "price": data["price"]}
        for name, data in supplier_data.items()
    ]

    return data_list, err_count


def main():
    try:
        with open("parse_errors_fast.log", "w", encoding="utf-8") as log_output:
            parsed_data, error_count = parse_excel_file("./supplier.xlsx", log_output)

            print(f"Данные успешно спарсены: {len(parsed_data)} строк")
            print(f"Ошибок при обработке: {error_count}")

            for row in parsed_data[:5]:
                print(row)

            # Сохраняем результат в JSON
            with open("parsed_data_fast.json", "w", encoding="utf-8") as f_out:
                json.dump(parsed_data, f_out, ensure_ascii=False, indent=2)

            # Выводим первые 50 символов лога
            log_output.seek(0)
            print("\nСообщения лога (первые 50 символов):")
            print(log_output.read(50))

    except FileNotFoundError:
        print("Ошибка: Файл 'supplier.xlsx' не найден.")
    except ValueError as ve:
        print(f"Ошибка валидации: {ve}")
    except Exception as e:
        print(f"Произошла неожиданная ошибка: {e}")


if __name__ == "__main__":
    main()
